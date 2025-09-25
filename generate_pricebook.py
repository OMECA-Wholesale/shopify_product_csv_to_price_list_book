import json
import os
import sys
import pandas as pd
from datetime import datetime
import requests
from io import BytesIO
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from typing import Dict, List, Optional
import glob

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from lib.product_extractor import ProductExtractor
from lib.translation_extractor import TranslationExtractor

class PriceBookGenerator:
    def __init__(self, config_path: str = "config.json"):
        self.config = self.load_config(config_path)
        self.product_extractor = None
        self.translation_extractor = None
        self.wb = None
        self.ws = None
        self.temp_images = []  # Store temporary image paths

    def load_config(self, config_path: str) -> Dict:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def initialize_extractors(self):
        product_csv_files = glob.glob("inputs/shopify_product_csv/*.csv")
        if not product_csv_files:
            raise FileNotFoundError("No product CSV files found in inputs/shopify_product_csv/")

        self.product_extractor = ProductExtractor(product_csv_files[0])
        self.product_extractor.load_data()
        self.product_extractor.extract_products()

        translation_csv_files = glob.glob("inputs/shopify_translate_csv/*.csv")
        if translation_csv_files:
            self.translation_extractor = TranslationExtractor(translation_csv_files[0])
            self.translation_extractor.load_data()
            self.translation_extractor.extract_translations()

    def create_workbook(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Price List"

    def add_header(self, start_row: int = 1) -> int:
        # Company name styling
        company_font = Font(size=24, bold=True, color="1F4788")
        info_font = Font(size=11, color="4A4A4A")

        # Check if logo exists and add it
        logo_path = self.config.get('logo', '')
        logo_added = False
        if logo_path and os.path.exists(logo_path):
            try:
                # Load and resize logo
                logo_img = Image.open(logo_path)

                # Resize logo to fit header (max height 60px)
                max_height = 60
                aspect_ratio = logo_img.width / logo_img.height
                new_height = min(logo_img.height, max_height)
                new_width = int(new_height * aspect_ratio)

                logo_img.thumbnail((new_width, new_height), Image.Resampling.LANCZOS)

                # Save resized logo temporarily
                temp_logo = "temp/logo_header.png"
                os.makedirs("temp", exist_ok=True)
                logo_img.save(temp_logo)
                self.temp_images.append(temp_logo)

                # Add logo to Excel
                xl_logo = XLImage(temp_logo)
                xl_logo.height = new_height
                xl_logo.width = new_width

                # Position logo in column A
                self.ws.add_image(xl_logo, f'A{start_row}')
                logo_added = True

                # Adjust row height for logo
                self.ws.row_dimensions[start_row].height = max(45, new_height * 0.75)
            except Exception as e:
                print(f"Could not add logo: {e}")
                self.ws.row_dimensions[start_row].height = 35
        else:
            self.ws.row_dimensions[start_row].height = 35

        # Company Name - always use full width for centering
        self.ws.merge_cells(f'A{start_row}:E{start_row}')
        cell = self.ws[f'A{start_row}']
        cell.value = self.config.get('company_name', 'Company Name')
        cell.font = company_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add a subtle background
        for col in range(1, 6):
            self.ws.cell(row=start_row, column=col).fill = PatternFill(
                start_color="F0F4F8", end_color="F0F4F8", fill_type="solid"
            )

        # Contact info
        info_row = start_row + 1
        self.ws.row_dimensions[info_row].height = 20
        self.ws.merge_cells(f'A{info_row}:E{info_row}')
        contact_info = f"📞 {self.config.get('phone', '')}  "
        if self.config.get('email'):
            contact_info += f"✉️ {self.config.get('email', '')}"
        if self.config.get('website'):
            contact_info += f"  🌐 {self.config.get('website', '')}"
        self.ws[f'A{info_row}'] = contact_info
        self.ws[f'A{info_row}'].font = info_font
        self.ws[f'A{info_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # Address
        if self.config.get('address'):
            addr_row = info_row + 1
            self.ws.row_dimensions[addr_row].height = 20
            self.ws.merge_cells(f'A{addr_row}:E{addr_row}')
            self.ws[f'A{addr_row}'] = f"📍 {self.config.get('address', '')}"
            self.ws[f'A{addr_row}'].font = info_font
            self.ws[f'A{addr_row}'].alignment = Alignment(horizontal='center', vertical='center')

            # Add a divider line
            divider_row = addr_row + 1
            self.ws.row_dimensions[divider_row].height = 5
            for col in range(1, 6):
                self.ws.cell(row=divider_row, column=col).fill = PatternFill(
                    start_color="1F4788", end_color="1F4788", fill_type="solid"
                )
            return divider_row + 2

        return info_row + 3

    def add_product_section(self, products: List[Dict], section_title: str, start_row: int) -> int:
        # Section title styling
        section_font = Font(size=16, bold=True, color="FFFFFF")
        header_font = Font(size=11, bold=True, color="FFFFFF")

        # Modern border style
        thin_border = Border(
            left=Side(style='thin', color="E0E0E0"),
            right=Side(style='thin', color="E0E0E0"),
            top=Side(style='thin', color="E0E0E0"),
            bottom=Side(style='thin', color="E0E0E0")
        )

        # Section header with gradient-like effect
        section_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        header_fill = PatternFill(start_color="4A6FA5", end_color="4A6FA5", fill_type="solid")

        # Section title
        self.ws.row_dimensions[start_row].height = 30
        self.ws.merge_cells(f'A{start_row}:E{start_row}')
        self.ws[f'A{start_row}'] = f"  {section_title.upper()}  "
        self.ws[f'A{start_row}'].font = section_font
        self.ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws[f'A{start_row}'].fill = section_fill

        for col in range(1, 6):
            self.ws.cell(row=start_row, column=col).border = thin_border

        current_row = start_row + 1

        if not products:
            return current_row + 1

        # Column headers
        self.ws.row_dimensions[current_row].height = 25
        headers = ['Image', 'SKU', 'Product Name', 'Variant', 'Wholesale Price']
        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
            cell.border = thin_border

        current_row += 1

        # Track row colors for alternating effect
        row_count = 0

        for product_idx, product in enumerate(products):
            # Store the starting row for this product
            product_start_row = current_row

            # Count variants to know how many rows to merge
            variants = product.get('variants', [{}])
            variant_count = len(variants)

            # We'll add the image after merging cells
            image_to_add = None
            if product.get('images'):
                image_url = product['images'][0].get('src', '')
                if image_url:
                    try:
                        response = requests.get(image_url, timeout=10)
                        if response.status_code == 200:
                            img = Image.open(BytesIO(response.content))

                            max_size = (100, 100)
                            img.thumbnail(max_size, Image.Resampling.LANCZOS)

                            # Create temp directory if it doesn't exist
                            os.makedirs("temp", exist_ok=True)

                            # Use unique filename with product handle to avoid conflicts
                            temp_path = f"temp/temp_img_{product.get('handle', 'unknown')}_{product_start_row}.png"
                            img.save(temp_path)
                            self.temp_images.append(temp_path)  # Track temp files

                            image_to_add = temp_path

                            # Don't delete here - will cleanup after save
                    except Exception as e:
                        print(f"Error loading image: {e}")

            # Get product name for merging across variants
            languages = self.config.get('target_language', ['default'])
            product_names = []

            for lang in languages:
                if lang == 'default':
                    product_names.append(product.get('title', ''))
                elif self.translation_extractor:
                    translated = self.translation_extractor.get_translated_title(
                        product.get('handle', ''),
                        lang
                    )
                    if translated:
                        product_names.append(translated)

            product_name_combined = '\n'.join(product_names)

            for variant_idx, variant in enumerate(variants):
                row_data = []

                # Image column
                row_data.append("")

                # SKU
                row_data.append(variant.get('sku', ''))

                # Variant string
                variant_parts = []
                option1 = variant.get('option1', '')
                if option1 and option1 != 'Default Title' and str(option1).lower() != 'nan':
                    variant_parts.append(str(option1))
                option2 = variant.get('option2', '')
                if option2 and str(option2).lower() != 'nan':
                    variant_parts.append(str(option2))
                option3 = variant.get('option3', '')
                if option3 and str(option3).lower() != 'nan':
                    variant_parts.append(str(option3))

                variant_str = ' / '.join(variant_parts) if variant_parts else ''

                # Price
                price = variant.get('price', 0)
                try:
                    price_val = float(price) if price else 0
                    row_data.append(f"${price_val:.2f}")
                except:
                    row_data.append(str(price))

                # Determine row color (alternating)
                if product_idx % 2 == 0:
                    row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                else:
                    row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

                # Define borders for product group
                # Thicker border for product boundaries
                thick_border = Border(
                    left=Side(style='medium', color="4A6FA5"),
                    right=Side(style='medium', color="4A6FA5"),
                    top=Side(style='medium', color="4A6FA5") if variant_idx == 0 else Side(style='thin', color="E0E0E0"),
                    bottom=Side(style='medium', color="4A6FA5") if variant_idx == variant_count - 1 else Side(style='thin', color="E0E0E0")
                )

                # Image column (A) - apply border
                image_cell = self.ws.cell(row=current_row, column=1)
                image_cell.border = thick_border
                image_cell.fill = row_fill

                # SKU column (B)
                sku_cell = self.ws.cell(row=current_row, column=2, value=row_data[1])
                sku_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sku_cell.border = thick_border
                sku_cell.fill = row_fill
                sku_cell.font = Font(color="2C3E50", size=14)

                # Variant column (D)
                cell = self.ws.cell(row=current_row, column=4, value=variant_str)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thick_border
                cell.fill = row_fill
                cell.font = Font(color="2C3E50", size=14)

                # Price column (E)
                cell = self.ws.cell(row=current_row, column=5, value=row_data[-1])
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = thick_border
                cell.fill = row_fill
                cell.font = Font(color="2C3E50", size=14)

                # Set row height based on variant count
                if variant_count == 1:
                    # Single variant - standard height
                    self.ws.row_dimensions[current_row].height = 85
                else:
                    # Multiple variants - calculate reduced height
                    # Total height should be around 100-120 for image display
                    # Distribute height across variants
                    if variant_idx == 0:
                        # First row gets slightly more height
                        self.ws.row_dimensions[current_row].height = max(30, min(50, 100 / variant_count + 10))
                    else:
                        # Other rows get proportional height
                        self.ws.row_dimensions[current_row].height = max(25, min(40, 100 / variant_count))

                current_row += 1

            # After all variants, handle merged cells
            if variant_count > 1:
                # Merge cells for image
                self.ws.merge_cells(f'A{product_start_row}:A{current_row - 1}')

                # Set alignment for merged image cell
                image_cell = self.ws[f'A{product_start_row}']
                image_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Merge cells for product name
                self.ws.merge_cells(f'C{product_start_row}:C{current_row - 1}')

                # Apply borders to all cells in the merged ranges
                for row_idx in range(product_start_row, current_row):
                    # Image column borders
                    image_cell = self.ws.cell(row=row_idx, column=1)
                    image_cell.border = Border(
                        left=Side(style='medium', color="4A6FA5"),
                        right=Side(style='medium', color="4A6FA5"),
                        top=Side(style='medium', color="4A6FA5") if row_idx == product_start_row else Side(style='thin', color="E0E0E0"),
                        bottom=Side(style='medium', color="4A6FA5") if row_idx == current_row - 1 else Side(style='thin', color="E0E0E0")
                    )
                    image_cell.fill = row_fill

                    # Product name column borders
                    name_cell = self.ws.cell(row=row_idx, column=3)
                    name_cell.border = Border(
                        left=Side(style='medium', color="4A6FA5"),
                        right=Side(style='medium', color="4A6FA5"),
                        top=Side(style='medium', color="4A6FA5") if row_idx == product_start_row else Side(style='thin', color="E0E0E0"),
                        bottom=Side(style='medium', color="4A6FA5") if row_idx == current_row - 1 else Side(style='thin', color="E0E0E0")
                    )
                    name_cell.fill = row_fill

            # Add image to merged cell (or single cell) - AFTER merging is complete
            if image_to_add:
                xl_img = XLImage(image_to_add)

                # Set image size smaller to fit better
                img_width = 100
                img_height = 100
                xl_img.width = img_width
                xl_img.height = img_height

                # Simply add the image to the first cell of the merged range
                # Excel will handle it within the merged cell
                self.ws.add_image(xl_img, f'A{product_start_row}')

                # For centering, we rely on the smaller size and let Excel handle the positioning
                # The image will appear in the merged cell area

            # Write product name to the first cell (merged or not)
            product_name_cell = self.ws[f'C{product_start_row}']
            product_name_cell.value = product_name_combined
            product_name_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            product_name_cell.font = Font(color="2C3E50", size=14)

            # If only one variant, apply borders and alignment to single cells
            if variant_count == 1:
                # Image cell border and alignment
                image_cell = self.ws[f'A{product_start_row}']
                image_cell.alignment = Alignment(horizontal='center', vertical='center')
                image_cell.border = Border(
                    left=Side(style='medium', color="4A6FA5"),
                    right=Side(style='medium', color="4A6FA5"),
                    top=Side(style='medium', color="4A6FA5"),
                    bottom=Side(style='medium', color="4A6FA5")
                )
                image_cell.fill = row_fill

                # Product name cell border
                product_name_cell.border = Border(
                    left=Side(style='medium', color="4A6FA5"),
                    right=Side(style='medium', color="4A6FA5"),
                    top=Side(style='medium', color="4A6FA5"),
                    bottom=Side(style='medium', color="4A6FA5")
                )
                product_name_cell.fill = row_fill

            # Add spacing between different products
            if product_idx < len(products) - 1:
                # Add an empty row for spacing
                self.ws.row_dimensions[current_row].height = 10
                current_row += 1

        return current_row + 2

    def set_column_widths(self):
        column_widths = {
            'A': 15,   # Image
            'B': 15,   # SKU
            'C': 45,   # Product Name
            'D': 20,   # Variant
            'E': 18    # Wholesale Price
        }

        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width

        # Add print settings for better output
        self.ws.print_options.horizontalCentered = True
        self.ws.print_options.verticalCentered = False
        self.ws.page_setup.orientation = 'landscape'
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = False

    def generate(self):
        print("Initializing extractors...")
        self.initialize_extractors()

        print("Creating workbook...")
        self.create_workbook()

        print("Adding header...")
        current_row = self.add_header()

        print("Grouping products by tag...")
        grouped_products = self.product_extractor.group_products_by_tag()

        # Filter by target_tag if specified
        target_tags = self.config.get('target_tag', [])
        if target_tags and isinstance(target_tags, list) and len(target_tags) > 0:
            filtered_groups = {}
            for tag in target_tags:
                if tag in grouped_products:
                    filtered_groups[tag] = grouped_products[tag]
            if filtered_groups:  # Only use filtered groups if we found matches
                grouped_products = filtered_groups
            else:
                print(f"Warning: No products found with tags: {target_tags}")

        print(f"Found {len(grouped_products)} product groups")

        for tag, products in grouped_products.items():
            print(f"Adding section: {tag} with {len(products)} products")
            current_row = self.add_product_section(products, tag, current_row)

        print("Setting column widths...")
        self.set_column_widths()

        os.makedirs("outputs", exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"outputs/pricebook_{timestamp}.xlsx"

        print(f"Saving to {output_file}...")
        self.wb.save(output_file)
        print(f"Price book generated successfully: {output_file}")

        # Cleanup temporary images after save
        self.cleanup_temp_images()

        return output_file

    def cleanup_temp_images(self):
        """Remove temporary image files after workbook is saved"""
        for temp_path in self.temp_images:
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except Exception as e:
                print(f"Warning: Could not remove temp file {temp_path}: {e}")
        self.temp_images = []

def main():
    generator = PriceBookGenerator()
    generator.generate()

if __name__ == "__main__":
    main()